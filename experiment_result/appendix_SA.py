import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import pylab
import sys
import random
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
import pandas as pd
from pandas import DataFrame

letters = string.ascii_uppercase
letters_lower = string.ascii_lowercase
scenarios = ['HH','HL','LH','LL']
runs = 100
benchmark_no = 20

data = []
for ax_idx in range(4):
    '''import the data'''
    path = sys.path[0]+'\\S_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_tardy_rate = load_workbook(path)['tardy rate']
    data_maximum = load_workbook(path)['maximum']
    '''retrive the data'''
    name = []
    sum = []
    rate = []
    maxim = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append(np.mean([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        rate.append(np.mean([data_tardy_rate[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        maxim.append(np.mean([data_maximum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
    #print(len(avg),len(rate),len(maxim))
    data.append(sum)
    data.append(rate)
    data.append(maxim)

data.insert(0,name)
print(data)
data = np.array(data).transpose()
print(data.shape)


title = ['Rule'] + ['Sum','Tardy %','Maximum','Sum','Tardy %','Maximum','Sum','Tardy %','Maximum','Sum','Tardy %','Maximum']

df = DataFrame(data, columns=title)
address = sys.path[0]+'\\Appendix_SA.xlsx'
Excelwriter = pd.ExcelWriter(address,engine="xlsxwriter")

df.to_excel(Excelwriter, sheet_name='data', index=False)
Excelwriter.save()
