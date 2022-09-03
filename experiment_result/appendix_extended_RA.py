import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import pylab
import sys
import random
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
import pandas as pd
from pandas import DataFrame

letters = string.ascii_uppercase
letters_lower = string.ascii_lowercase
scenarios = ['HH','HL','LH','LL']
ops = [6,9]
runs = 100
benchmark_no = 3

data_3m_per_m = []
data_4m_per_m = []

for idx in range(4):
    '''import the data'''
    path = sys.path[0]+'\\ext_R_3m_per_wc_{}.xlsx'.format(scenarios[idx])
    data_sum = load_workbook(path)['sum']
    data_tardy_rate = load_workbook(path)['tardy rate']
    data_maximum = load_workbook(path)['maximum']
    '''retrive the data'''
    name = []
    sum = []
    rate = []
    maxim = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append(np.mean([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        rate.append(np.mean([data_tardy_rate[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        maxim.append(np.mean([data_maximum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
    #print(len(avg),len(rate),len(maxim))
    data_3m_per_m.append(name)
    data_3m_per_m.append(sum)
    data_3m_per_m.append(rate)
    data_3m_per_m.append(maxim)
data_3m_per_m = np.array(data_3m_per_m).transpose()

for idx in range(4):
    '''import the data'''
    path = sys.path[0]+'\\ext_R_4m_per_wc_{}.xlsx'.format(scenarios[idx])
    data_sum = load_workbook(path)['sum']
    data_tardy_rate = load_workbook(path)['tardy rate']
    data_maximum = load_workbook(path)['maximum']
    '''retrive the data'''
    name = []
    sum = []
    rate = []
    maxim = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append(np.mean([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        rate.append(np.mean([data_tardy_rate[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        maxim.append(np.mean([data_maximum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
    #print(len(avg),len(rate),len(maxim))
    data_4m_per_m.append(name)
    data_4m_per_m.append(sum)
    data_4m_per_m.append(rate)
    data_4m_per_m.append(maxim)
data_4m_per_m = np.array(data_4m_per_m).transpose()


title = ['Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum']
df1 = DataFrame(data_3m_per_m, columns=title)
df2 = DataFrame(data_4m_per_m, columns=title)
dflist = [df1,df2]
address = sys.path[0]+'\\Appendix_extended_RA.xlsx'
Excelwriter = pd.ExcelWriter(address,engine="xlsxwriter")
sheetnames = ['6ops','9ops']
for idx,df in enumerate(dflist):
    df.to_excel(Excelwriter, sheet_name=sheetnames[idx], index=False)
Excelwriter.save()
