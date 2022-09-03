import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import pylab
import sys
import random
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
import pandas as pd
from pandas import DataFrame

letters = string.ascii_uppercase
letters_lower = string.ascii_lowercase
scenarios = ['HH','HL','LH','LL']
ops = [6,9]
runs = 100
benchmark_no = 6

data_6ops = []
data_9ops = []
datalist = [data_6ops, data_9ops]

for idx in range(4):
    '''import the data'''
    path = sys.path[0]+'\\ext_S_6wc_{}.xlsx'.format(scenarios[idx])
    data_sum = load_workbook(path)['sum']
    data_tardy_rate = load_workbook(path)['tardy rate']
    data_maximum = load_workbook(path)['maximum']
    '''retrive the data'''
    name = []
    sum = []
    rate = []
    maxim = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append(np.mean([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        rate.append(np.mean([data_tardy_rate[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        maxim.append(np.mean([data_maximum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
    #print(len(avg),len(rate),len(maxim))
    data_6ops.append(name)
    data_6ops.append(sum)
    data_6ops.append(rate)
    data_6ops.append(maxim)
data_6ops = np.array(data_6ops).transpose()

for idx in range(4):
    '''import the data'''
    path = sys.path[0]+'\\ext_S_9wc_{}.xlsx'.format(scenarios[idx])
    data_sum = load_workbook(path)['sum']
    data_tardy_rate = load_workbook(path)['tardy rate']
    data_maximum = load_workbook(path)['maximum']
    '''retrive the data'''
    name = []
    sum = []
    rate = []
    maxim = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append(np.mean([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        rate.append(np.mean([data_tardy_rate[letters[idx] + str(i)].value for i in range(2,2+runs)]))
        maxim.append(np.mean([data_maximum[letters[idx] + str(i)].value for i in range(2,2+runs)]))
    #print(len(avg),len(rate),len(maxim))
    data_9ops.append(name)
    data_9ops.append(sum)
    data_9ops.append(rate)
    data_9ops.append(maxim)
data_9ops = np.array(data_9ops).transpose()


title = ['Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum','Rule','Sum','Tardy %','Maximum']
df1 = DataFrame(data_6ops, columns=title)
df2 = DataFrame(data_9ops, columns=title)
dflist = [df1,df2]
address = sys.path[0]+'\\Appendix_extended_SA.xlsx'
Excelwriter = pd.ExcelWriter(address,engine="xlsxwriter")
sheetnames = ['6ops','9ops']
for idx,df in enumerate(dflist):
    df.to_excel(Excelwriter, sheet_name=sheetnames[idx], index=False)
Excelwriter.save()
