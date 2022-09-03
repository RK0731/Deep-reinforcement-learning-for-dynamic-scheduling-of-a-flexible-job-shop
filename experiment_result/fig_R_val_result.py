import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel = ['({}.{}) {}','({}.{}) Win %, DRL(✖)','({}.{}) Win %, DRL(✔)']

letters = string.ascii_uppercase
letters_lower = string.ascii_lowercase
benchmark_no = 6
scenarios = ['HH','HL','LH','LL']
runs = 100
base = ['#029386','#f97306','#c2b709','#c20078','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffff14','#c44240','#b1916e']
addon = ['#0504aa']
color_list = base[:benchmark_no]+addon
bplot_color_list = ['w' for i in range(benchmark_no-1)]+addon

'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(10,8))
all_scenarios = gridspec.GridSpec(1, 1, figure=fig)
section = gridspec.GridSpecFromSubplotSpec(30, 40, subplot_spec = all_scenarios[0])
ax = {}
h_pos=[0,10,20,30]

for ax_idx in range(len(scenarios)):
    '''import the data'''
    path = sys.path[0]+'\\R_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_before_win_rate = load_workbook(path)['before win rate']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    #ax[ax_idx] = fig.add_subplot(section[bound*(ax_idx%2):bound*(ax_idx%2+1) , 10*(ax_idx%2):10%(ax_idx%2+1)])
    ax[ax_idx] = fig.add_subplot(section[:9 , h_pos[ax_idx]:h_pos[ax_idx]+9])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=-0.2, top=1)
    ax[ax_idx].set_xlim(left = -1, right=benchmark_no)
    '''retrive the data'''
    name = []
    sum = []
    # retrive the data
    for idx in range(benchmark_no+1):
        name.append(data_sum[letters[idx] + '1'].value)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
    name[-1] = r'$\mathbf{DRL-RA}$'



    '''common legend'''
    ax[100] = fig.add_subplot(section[27:, :])
    # common legend
    legend_elements = [Patch(facecolor=color_list[i],label=name[i]) for i in range(benchmark_no+1)]
    ax[100].legend(handles=legend_elements, fontsize=8, loc=2, ncol=3, )
    ax[100].axis('off')
    # annotations
    ax[100].set_xlim(0,100)
    ax[100].set_ylim(0,100)
    sign_x = 35
    ax[100].scatter(sign_x,60,marker="*",s=120,color='#fffd01',edgecolors='k')
    ax[100].text(sign_x+2,60,'Top two benchmark rules',va='center',fontsize=9)

    ax[100].scatter(sign_x,30,marker="X",s=70,color='r',edgecolors='k')
    ax[100].text(sign_x+2,30,'Lower-than-baseline performance',va='center',fontsize=9)

    sign_x = 70
    ax[100].text(sign_x,60,'DRL(✖)',va='center',fontsize=9, ha='center')
    ax[100].text(sign_x+5,60,'Result excluding DRL agents',va='center',fontsize=9)

    ax[100].text(sign_x,30,'DRL(✔)',va='center',fontsize=9, ha='center')
    ax[100].text(sign_x+5,30,'Result including DRL agents',va='center',fontsize=9)



    '''plot the data'''
    #print(sum)
    sum = 1 - sum / np.array(sum[0])
    #print(sum)
    name.pop(0) # drop the FIFO
    sum = np.delete(sum,0,axis=0)
    # create the plots
    #print(sum[:benchmark_no-1].mean(axis=1),sum.mean(axis=1),sum.mean(axis=1).argsort())
    tops = sum[:5].mean(axis=1).argsort()[-2:]
    bottoms = np.where(sum.mean(axis=1)<0)[0]
    drift = sum[:5].mean(axis=1)[tops].clip(-0.01,0)*10
    print(drift)
    #print(tops)
    x_position = np.arange(len(name))
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    ax[ax_idx].scatter(tops, np.zeros_like(tops)-drift,marker="*",s=120,color='#fffd01',edgecolors='k',zorder=5)
    ax[ax_idx].scatter(bottoms, np.zeros_like(bottoms),marker="X",s=70,color='r',edgecolors='k',zorder=5)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(-0.2,1.1,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    if ax_idx !=0:
        plt.setp(ax[ax_idx].get_yticklabels(), visible=0)
    if ax_idx==0:
        ax[ax_idx].set_ylabel('Normalised performance %', fontsize=10)
        plt.setp(ax[ax_idx].get_yticklabels(), visible=1)
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(), rotation=25, ha='right', rotation_mode="anchor", fontsize=7.5)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=9)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=benchmark_no, colors='k', linestyles='solid', linewidths=1)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=benchmark_no, colors='g', linestyles='--', linewidths=1, zorder=1)
    # fill
    ax[ax_idx].fill_between([-1,benchmark_no+1], [0,0], [1.2,1.2], color='c', alpha=0.15)
    # title
    ax[ax_idx].set_title(panel[0].format(letters_lower[ax_idx],1,scenarios[ax_idx]), fontsize=10)
    # legends
    legend_color = ['g']
    legend_line = ['--']
    legend_label = ['mean of DRL-RA']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(1)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=8, loc=2, ncol=2)



    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[12:18 , h_pos[ax_idx]+2:h_pos[ax_idx]+8])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_before_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_before_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3  else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, labels = label, labeldistance=1.15, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8.5))
    plt.setp(texts, ha='center', )
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=8.5, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment, **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12

    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[1].format(letters_lower[ax_idx],2,scenarios[ax_idx]), fontsize=10)
    #ax[ax_idx].set_xlim(-1.2,1.2)



    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[20:26 , h_pos[ax_idx]+2:h_pos[ax_idx]+8])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3 else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, explode=explode, labels = label, labeldistance=1.18, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8.5))
    plt.setp(texts, ha='center', )
    plt.setp(texts[-1], fontsize=10.5, fontweight='bold')
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=8.5, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})

        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment,  **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12
    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[2].format(letters_lower[ax_idx],3,scenarios[ax_idx]), fontsize=10)
    #ax[ax_idx].set_xlim(-1.2,1.2)



fig.subplots_adjust(top=0.95, bottom=0.1, hspace=0.5)
fig.savefig(sys.path[0]+"/R_result.png", dpi=600, bbox_inches='tight')
plt.show()
