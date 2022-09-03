import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel = ['({}.{}) {}','({}.{}) Win %\n DRL(✖)','({}.{}) Win %\n DRL(✔)']

letters = string.ascii_uppercase
letters_lower = string.ascii_lowercase
benchmark_no = 6
scenarios = ['HH','HL','LH','LL']
runs = 100
base = ['#7e1e9c','#15b01a','#0343df','#ff81c0','#653700','#e50000','#fac205','#029386','#f97306','#c2b709','#00ffff','#00035b','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffff14']
correspondence = ['FIFO','ATC','AVPRO','COVERT','CR','EDD','LWKR','MDD','MOD','MON','MS','NPT','SPT','WINQ','LWKRSPT','LWKRMOD','PTWINQ','PTWINQS','DPTLWKRS','DPTWINQNPT']
addon = ['#cb416b']
bplot_color_list = ['w' for i in range(benchmark_no-1)]+addon
grand_legend_name = []
grand_legend_color = []

'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(11,8.5))
all_scenarios = gridspec.GridSpec(1, 1, figure=fig)
section = gridspec.GridSpecFromSubplotSpec(67, 80, subplot_spec = all_scenarios[0])
ax = {}
h_pos=[0,20,40,60]

for ax_idx in range(len(scenarios)):
    '''import the data'''
    path = sys.path[0]+'\\ext_S_6wc_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_before_win_rate = load_workbook(path)['before win rate']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    #ax[ax_idx] = fig.add_subplot(section[bound*(ax_idx%2):bound*(ax_idx%2+1) , 10*(ax_idx%2):10%(ax_idx%2+1)])
    ax[ax_idx] = fig.add_subplot(section[3:23 , h_pos[ax_idx]:h_pos[ax_idx]+10])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0, top=0.8)
    ax[ax_idx].set_xlim(left = -0.5, right=benchmark_no-0.5)
    '''retrive the data'''
    name = []
    sum = []
    color_list = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
        if idx!=benchmark_no:
            color_list.append(base[correspondence.index(x)])
    color_list += addon
    name[-1] = r'$\mathbf{DRL-SA}$'
    grand_legend_name += name
    grand_legend_color += color_list
    #print(sum)
    sum = 1 - sum / np.array(sum[0])
    #print(sum)
    name.pop(0) # drop the FIFO
    sum = np.delete(sum,0,axis=0)
    # create the plots
    #print(sum[:benchmark_no-1].mean(axis=1),sum.mean(axis=1),sum.mean(axis=1).argsort())
    tops = sum[:2].mean(axis=1).argsort()[-2:]
    bottoms = np.where(sum.mean(axis=1)<0)[0]
    #print(tops)
    x_position = np.arange(len(name))
    '''plot the data'''
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    #ax[ax_idx].scatter(tops, np.zeros_like(tops),marker="*",s=120,color='#fffd01',edgecolors='k',zorder=5)
    ax[ax_idx].scatter(bottoms, np.zeros_like(bottoms),marker="X",s=70,color='r',edgecolors='k',zorder=5)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(0,0.9,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    #if ax_idx !=0:
        #plt.setp(ax[ax_idx].get_yticklabels(), visible=0)
    if ax_idx==0:
        ax[ax_idx].set_ylabel('Normalised performance %', fontsize=10)
        #plt.setp(ax[ax_idx].get_yticklabels(), visible=1)
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(),  ha='right', rotation=40, rotation_mode="anchor", fontsize=7.5)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=7)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=benchmark_no, colors='k', linestyles='solid', linewidths=1)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=benchmark_no, colors='g', linestyles='--', linewidths=1, zorder=1)
    # fill
    ax[ax_idx].fill_between([-1,benchmark_no+1], [0,0], [1.2,1.2], color='r', alpha=0.05)
    # title
    ax[ax_idx].set_title(panel[0].format(letters_lower[ax_idx],1,scenarios[ax_idx]), fontsize=8)
    # legends
    legend_color = ['g']
    legend_line = ['--']
    legend_label = ['mean of DRL-SA']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(1)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=6, loc=1, ncol=2)



    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[4:13 , h_pos[ax_idx]+11:h_pos[ax_idx]+17])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_before_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_before_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3  else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, labels = label, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=7))
    plt.setp(texts, ha='center', )
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=7, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment, **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12

    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[1].format(letters_lower[ax_idx],2,scenarios[ax_idx]), fontsize=8)
    #ax[ax_idx].set_xlim(-1.2,1.2)

    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[16:25 , h_pos[ax_idx]+11:h_pos[ax_idx]+17])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3 else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, explode=explode, labels = label, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=7))
    plt.setp(texts, ha='center', )
    plt.setp(texts[-1], fontsize=8.2, fontweight='bold')
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=7, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})

        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment, **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12
    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[2].format(letters_lower[ax_idx],3,scenarios[ax_idx]), fontsize=8)
    #ax[ax_idx].set_xlim(-1.2,1.2)


# 30:50 31:40 43 52

for ax_idx in range(len(scenarios)):
    '''import the data'''
    path = sys.path[0]+'\\ext_S_9wc_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_before_win_rate = load_workbook(path)['before win rate']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    #ax[ax_idx] = fig.add_subplot(section[bound*(ax_idx%2):bound*(ax_idx%2+1) , 10*(ax_idx%2):10%(ax_idx%2+1)])
    ax[ax_idx] = fig.add_subplot(section[35:55 , h_pos[ax_idx]:h_pos[ax_idx]+10])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0, top=0.8)
    ax[ax_idx].set_xlim(left = -0.5, right=benchmark_no-0.5)
    '''retrive the data'''
    name = []
    sum = []
    color_list = []
    # retrive the data
    for idx in range(benchmark_no+1):
        x = data_sum[letters[idx] + '1'].value
        name.append(x)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
        if idx!=benchmark_no:
            color_list.append(base[correspondence.index(x)])
    color_list += addon
    name[-1] = r'$\mathbf{DRL-SA}$'
    grand_legend_name += name
    grand_legend_color += color_list
    #print(sum)
    sum = 1 - sum / np.array(sum[0])
    #print(sum)
    name.pop(0) # drop the FIFO
    sum = np.delete(sum,0,axis=0)
    # create the plots
    #print(sum[:benchmark_no-1].mean(axis=1),sum.mean(axis=1),sum.mean(axis=1).argsort())
    tops = sum[:2].mean(axis=1).argsort()[-2:]
    bottoms = np.where(sum.mean(axis=1)<0)[0]
    #print(tops)
    x_position = np.arange(len(name))
    '''plot the data'''
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    #ax[ax_idx].scatter(tops, np.zeros_like(tops),marker="*",s=120,color='#fffd01',edgecolors='k',zorder=5)
    ax[ax_idx].scatter(bottoms, np.zeros_like(bottoms),marker="X",s=70,color='r',edgecolors='k',zorder=5)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(0,0.9,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    #if ax_idx !=0:
        #plt.setp(ax[ax_idx].get_yticklabels(), visible=0)
    if ax_idx==0:
        ax[ax_idx].set_ylabel('Normalised performance %', fontsize=10)
        #plt.setp(ax[ax_idx].get_yticklabels(), visible=1)
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(),  ha='right', rotation=40, rotation_mode="anchor", fontsize=7.5)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=7)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=benchmark_no, colors='k', linestyles='solid', linewidths=1)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=benchmark_no, colors='g', linestyles='--', linewidths=1, zorder=1)
    # fill
    ax[ax_idx].fill_between([-1,benchmark_no+1], [0,0], [1.2,1.2], color='r', alpha=0.05)
    # title
    ax[ax_idx].set_title(panel[0].format(letters_lower[ax_idx+4],1,scenarios[ax_idx]), fontsize=8)
    # legends
    legend_color = ['g']
    legend_line = ['--']
    legend_label = ['mean of DRL-SA']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(1)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=6, loc=1, ncol=2)



    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[36:45 , h_pos[ax_idx]+11:h_pos[ax_idx]+17])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_before_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_before_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3  else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, labels = label, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=7))
    plt.setp(texts, ha='center', )
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=7, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment, **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12

    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[1].format(letters_lower[ax_idx+4],2,scenarios[ax_idx]), fontsize=8)
    #ax[ax_idx].set_xlim(-1.2,1.2)

    '''and the pie chart'''
    ax[ax_idx] = fig.add_subplot(section[48:57 , h_pos[ax_idx]+11:h_pos[ax_idx]+17])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no+1):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3 else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx].pie(win_rate, explode=explode, labels = label, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=7))
    plt.setp(texts, ha='center', )
    plt.setp(texts[-1], fontsize=8.2, fontweight='bold')
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=7, )
    pre=1
    next = [1.2,1.35]
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})

        if annotates[i]:
            #print(annotates[i],1.3*y,pre)
            if np.abs(pre-1.2*y) > 0.15 or pre==1:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.2*y),horizontalalignment=horizontalalignment, **kw)
                pre = 1.2*y
            else:
                ax[ax_idx].annotate(annotates[i], xy=(x, y), xytext=(next[i%2]*np.sign(x), pre+np.sign(x)*0.12),horizontalalignment=horizontalalignment, **kw)
                pre = pre+np.sign(x)*0.12
    ax[ax_idx].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx].set_title(panel[2].format(letters_lower[ax_idx+4],3,scenarios[ax_idx]), fontsize=8)
    #ax[ax_idx].set_xlim(-1.2,1.2)


'''common legend'''
ax[100] = fig.add_subplot(section[62:, :])
# common legend
name = list(set(grand_legend_name))
for i in correspondence.copy():
    #print(i)
    if i not in name:
        idx = correspondence.index(i)
        del correspondence[idx]
        del base[idx]
base += addon
correspondence += [r'$\mathbf{DRL-SA}$']
#print(len(correspondence),len(name))
legend_elements = [Patch(facecolor=base[i],label=correspondence[i]) for i in range(len(correspondence))]
legend1 = ax[100].legend(handles=legend_elements, fontsize=8, loc=2, ncol=5)
ax[100].axis('off')
ax[100].add_artist(legend1)
# annotations
ax[100].set_xlim(0,100)
ax[100].set_ylim(0,100)
sign_x = 67
ax[100].text(sign_x,50,'DRL(✖)',va='center',fontsize=9, ha='center')
ax[100].text(sign_x+5,50,'Result excluding DRL agents',va='center',fontsize=9)

ax[100].text(sign_x,10,'DRL(✔)',va='center',fontsize=9, ha='center')
ax[100].text(sign_x+5,10,'Result including DRL agents',va='center',fontsize=9)



'''common title'''
ax[11] = fig.add_subplot(section[0:, :])
ax[11].set_title('(1) 6-operation test')
ax[11].axis('off')
ax[12] = fig.add_subplot(section[32:, :])
ax[12].set_title('(2) 9-operation test ')
ax[12].axis('off')


fig.subplots_adjust(top=0.95, bottom=0.1, hspace=0.5)
fig.savefig(sys.path[0]+"/ext_S_result.png", dpi=600, bbox_inches='tight')
plt.show()
