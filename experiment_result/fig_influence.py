import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data_decision_influence
name_col = 0
ones_col = 1
twos_col = 2
threes_col = 3
fours_col = 4
aboves_col = 5
panels = []
letters = string.ascii_uppercase
lower_letters = string.ascii_lowercase
color_list = ['w','#d1ffbd','#39ad48','#137e6d','k']
edge_list =  ['k','#d1ffbd','#39ad48','#137e6d','k']
#color_list = ['w','#d0fefe','#a2cffe','#0165fc','k']
#edge_list =  ['k','#d0fefe','#a2cffe','#0165fc','k']
hatch_list = ['//']+['w' for i in range(4)]
legend_list = ['passive decision','two jobs','three jobs', 'four jobs', 'five jobs and above']

scenarios = [70,80,90]
scenario_no = 3
runs = 50
name = ['3 wc/6 m','4 wc/12 m','5 wc/20 m']

# figure and subplots
ax={}
fig = plt.figure(figsize=(11,8.5))

#ax_legend = fig.add_subplot(111, visible=False)

''' plot the decision influence'''
for ax_idx, sce in enumerate(scenarios): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx+1)
    '''import the data'''
    path = sys.path[0]+'\\decision_influence_{}.xlsx'.format(sce)
    data_decision_influence = load_workbook(path)['active_decision']
    # retrive the data_decision_influence
    for idx in range(scenario_no):
        bars = [data_decision_influence[letters[i+1] + str(idx+2)].value for i in range(5)]
        #print(bars)
        bottom = np.cumsum(bars)
        bottom = np.delete(bottom,-1)
        bottom = np.insert(bottom,0,0)
        #print(bottom)
        # x position of bars
        x_position = idx
        # limit
        y_range = 1.0
        # plot the bars
        for i in range(5):
            ax[ax_idx].bar(x_position, bars[i], 0.3,  bottom=bottom[i], color=color_list[i], hatch=hatch_list[i], edgecolor=edge_list[i], align = 'center', zorder=3,)
    ax[ax_idx].set_ylim(0,y_range)
    ax[ax_idx].set_xlim(-0.5,2.5)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name, rotation=15, ha='right', rotation_mode="anchor", fontsize=9)
    ax[ax_idx].set_yticks(np.arange(0, 1.05, 0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    plt.setp(ax[ax_idx].get_xticklabels(), fontsize=9)
    plt.setp(ax[0].get_yticklabels(), visible=True)
    # labels
    ax[ax_idx].set_title('({}) UTIL.={}%'.format(lower_letters[ax_idx],scenarios[ax_idx]), fontsize=10)
    ax[0].set_ylabel('(1) Ratio of decisions %', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )

''' plot the tardy rate'''
for ax_idx, sce in enumerate(scenarios): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx+5)
    '''import the data'''
    path = sys.path[0]+'\\decision_influence_{}.xlsx'.format(sce)
    data_tardy_rate = load_workbook(path)['tardy rate']
    t_rate = []
    x_position = np.arange(3)
    for col in range(scenario_no):
        t_rate.append([data_tardy_rate[letters[col] + str(idx)].value for idx in range(2,2+runs)])
    ax[ax_idx].boxplot(t_rate, positions=x_position, showmeans=True, meanline=True, notch=True, zorder=3,)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name, rotation=15, ha='right', rotation_mode="anchor", fontsize=9)
    ax[ax_idx].set_yticks(np.arange(0, 1.05, 0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[0].set_ylabel('(2) Tardy Rate %', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )

''' plot the tardiness'''
for ax_idx, sce in enumerate(scenarios): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx+9)
    '''import the data'''
    path = sys.path[0]+'\\decision_influence_{}.xlsx'.format(sce)
    data_tardiness = load_workbook(path)['avg tardiness']
    t_rate = []
    x_position = np.arange(3)
    for col in range(scenario_no):
        t_rate.append([data_tardiness[letters[col] + str(idx)].value for idx in range(2,2+runs)])
    #t_rate = np.array(t_rate)
    ax[ax_idx].boxplot(t_rate, positions=x_position, showmeans=True, meanline=True, notch=True, zorder=3,)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name, rotation=15, ha='right', rotation_mode="anchor", fontsize=9)
    ax[ax_idx].set_yticks(np.arange(0, 110, 10))
    ax[ax_idx].set_ylim(bottom=0,top=100)
    ax[0].set_ylabel('(3) Average Tardiness', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )


'''common legends'''
ax_legend_0 = fig.add_subplot(3,4,4)
color_list.reverse()
edge_list.reverse()
hatch_list.reverse()
legend_list.reverse()
legend_elements = [Patch(facecolor=color_list[i], edgecolor=edge_list[i], hatch=hatch_list[i], label=legend_list[i]) for i in range(5)]

legend_elements.insert(0,Patch(facecolor='w', edgecolor='w', hatch='None', label='Only for SA'))

ax_legend_0.legend(handles=legend_elements, fontsize=10, loc=7)
ax_legend_0.axis('off')
# boxplot legend
ax_legend_1 = fig.add_subplot(2,4,8)
dummy = np.array([0.5,2,2,2,3,4,5,7,8,9,6,7,4,5,6,3,5,1,2,4,13])
ax_legend_1.set_xlim(0.9,1.8)
ax_legend_1.set_ylim(-6,13.5)
ax_legend_1.set_xticks(np.arange(1,2,0.5))
ax_legend_1.boxplot(dummy,showmeans=True, meanline=True, notch=True)
ax_legend_1.text(1.1,dummy.mean()-0.15,'mean',color='g')
ax_legend_1.text(1.1,np.median(dummy)-0.55,'median',color='tab:orange')
ax_legend_1.text(1.1,np.percentile(dummy,75),'75th percentile', verticalalignment='center')
ax_legend_1.text(1.1,np.percentile(dummy,25),'25th percentile', verticalalignment='center')
ax_legend_1.text(1.1,9,'Maximum', verticalalignment='center')
ax_legend_1.text(1.1,dummy.min(),'Minimum', verticalalignment='center')
ax_legend_1.text(1.1,dummy.max(),'Outlier', verticalalignment='center')
ax_legend_1.axis('off')
#ax_legend_2.axis('off')
#ax[7].set_xlabel('Number of Machines', fontsize=10)

fig.subplots_adjust(top=0.9, bottom=0.1, right=0.9, wspace=0.25, hspace=0.25)
#fig.savefig(sys.path[0]+"/Thesis_influence_tardy_rate.png", dpi=600, bbox_inches='tight')
plt.show()
